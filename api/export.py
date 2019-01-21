from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file

from db.connection import Connect
from datetime import datetime
from db.model import *
import json


def str2bool(str):
  if (str.upper() == "TRUE"):
    return True
  else:
    return False


class Export():
        def __init__(self):
                self.conn = Connect()
                self.session = self.conn.session()

        def __del__(self):
                self.session.close()

        def createSheetTanimlar(self, wb, model, title_, cid_):
                try:
                        if (model==ModelProfiller):
                            ws = wb.active
                            ws.title = title_
                        else:
                            ws = wb.create_sheet(title=title_)

                        # TITLE
                        ws['A1'] = "Kodu"
                        # ws['A1'].font = Font(bold=True)
                        ws['B1'] = title_

                        if (model == ModelSistemler):
                           ws['C1'] = "Dahili"
                           data = self.session.query(model.pidm, model.name, model.local).filter_by(cid=cid_)
                        elif (model == ModelUlkeler):
                           ws['C1'] = "Tel Kodu"
                           ws['D1'] = "Güvenli Ülke"
                           data = self.session.query(model.pidm, model.name, model.phone_area, model.secure).filter_by(cid=cid_)
                        else:
                           data = self.session.query(model.pidm, model.name).filter_by(cid=cid_)

                        data = data.order_by(model.name)

                        i=2
                        for row in data:
                                ws['A'+str(i)] = row.pidm
                                ws['B'+str(i)] = row.name
                                if (model == ModelSistemler):
                                        ws['C'+str(i)] = row.local
                                elif (model == ModelUlkeler):
                                        ws['C'+str(i)] = row.phone_area
                                        ws['D'+str(i)] = row.secure

                                i=i+1

                        return ws
                except Exception:
                        print("export.py > createSheetTanimlar Exception!")
                        return None

        def createSheetKVProfil(self, wb, cid_):
                try:
                        ws = wb.create_sheet("KV Profiller")

                        # TITLE
                        ws['A1'] = "Kodu"
                        ws['B1'] = "Profil"
                        ws['C1'] = "Birim"
                        ws['D1'] = "Veri Kategorileri"

                        ws['A1'].font = Font(bold=True)
                        ws['B1'].font = Font(bold=True)
                        ws['C1'].font = Font(bold=True)
                        ws['D1'].font = Font(bold=True)

                        # model=ViewKVProfilModel
                        model=ModelAnaveriler

                        data = self.session.query(model.pidm, model.profil_name, model.birim_name, model.data).filter_by(cid=cid_)
                        data = data.order_by(model.profil_name, model.birim_name)

                        i = 2
                        for row in data:
                                ws['A'+str(i)] = row.pidm
                                ws['B'+str(i)] = row.profil_name
                                ws['C'+str(i)] = row.birim_name
                                ws['D'+str(i)] = json.dumps(row.data)
                                i = i+1

                        return ws

                except Exception:
                        print("export.py > createSheetKVProfil Exception!")
                        return None

        def createSheetKVPaylasim(self, wb, cid_):
                try:
                        ws = wb.create_sheet("KV Paylaşımları")

                        # TITLE
                        ws['A1'] = "Kodu"
                        ws['B1'] = "Birim"
                        ws['C1'] = "Veri Kategorisi"
                        ws['D1'] = "Paylaşılan Kurum"
                        ws['E1'] = "İşleme Amaçları"
                        ws['F1'] = "Paylaşım Amaçları"
                        ws['G1'] = "Paylaşım Şekilleri"

                        ws['A1'].font = Font(bold=True)
                        ws['B1'].font = Font(bold=True)
                        ws['C1'].font = Font(bold=True)
                        ws['D1'].font = Font(bold=True)
                        ws['E1'].font = Font(bold=True)
                        ws['F1'].font = Font(bold=True)
                        ws['G1'].font = Font(bold=True)

                        # model=ViewKVPaylasimModel
                        model=ModelViewAktarimlar

                        data = self.session.query(model.pidm, model.birim_name, model.kv_name, model.kurum_name, model.islemeamaclari_data, model.paylasimamaclari_data, model.paylasimsekilleri_data).filter_by(cid=cid_)
                        data = data.order_by(model.birim_name, model.kv_name, model.kurum_name)

                        i = 2
                        for row in data:
                                ws['A'+str(i)] = row.pidm
                                ws['B'+str(i)] = row.birim_name
                                ws['C'+str(i)] = row.kv_name
                                ws['D'+str(i)] = row.kurum_name
                                ws['E'+str(i)] = json.dumps(row.islemeamaclari_data)
                                ws['F'+str(i)] = json.dumps(row.paylasimamaclari_data)
                                ws['G'+str(i)] = json.dumps(row.paylasimsekilleri_data)
                                i = i+1

                        return ws

                except Exception:
                        print("export.py > createSheetKVPaylasim Exception!")
                        return None

        def createSheetKVAnaveri(self, wb, cid_):
                try:
                        ws = wb.create_sheet("KV Ana Verileri")

                        # TITLE
                        ws['A1'] = "Kodu"
                        ws['B1'] = "Birim"
                        ws['C1'] = "Veri Kategorisi"
                        ws['D1'] = "Sure"
                        ws['E1'] = "Güvenli Ülkeler"
                        ws['F1'] = "Kanallar"
                        ws['G1'] = "Dokumanlar"
                        ws['H1'] = "Sistemler"
                        ws['I1'] = "Dayanaklar"
                        ws['J1'] = "Ortamlar"
                        ws['K1'] = "Tedbirler"

                        ws['A1'].font = Font(bold=True)
                        ws['B1'].font = Font(bold=True)
                        ws['C1'].font = Font(bold=True)
                        ws['D1'].font = Font(bold=True)
                        ws['E1'].font = Font(bold=True)
                        ws['F1'].font = Font(bold=True)
                        ws['G1'].font = Font(bold=True)
                        ws['H1'].font = Font(bold=True)
                        ws['I1'].font = Font(bold=True)
                        ws['J1'].font = Font(bold=True)
                        ws['K1'].font = Font(bold=True)

                        # model=ViewKVAnaveriModel
                        model=ModelViewAnaveriler

                        data = self.session.query(model.pidm, model.birim_name, model.kv_name, model.sure_name, model.ulkeler_data, model.kanallar_data, model.dokumanlar_data, model.sistemler_data, model.dayanaklar_data, model.ortamlar_data, model.tedbirler_data).filter_by(cid=cid_)
                        data = data.order_by(model.birim_name, model.kv_name)

                        i = 2
                        for row in data:
                                ws['A'+str(i)] = row.pidm
                                ws['B'+str(i)] = row.birim_name
                                ws['C'+str(i)] = row.kv_name
                                ws['D'+str(i)] = row.sure_name
                                ws['E'+str(i)] = json.dumps(row.ulkeler_data)
                                ws['F'+str(i)] = json.dumps(row.kanallar_data)
                                ws['G'+str(i)] = json.dumps(row.dokumanlar_data)
                                ws['H'+str(i)] = json.dumps(row.sistemler_data)
                                ws['I'+str(i)] = json.dumps(row.dayanaklar_data)
                                ws['J'+str(i)] = json.dumps(row.ortamlar_data)
                                ws['K'+str(i)] = json.dumps(row.tedbirler_data)
                                i = i+1

                        return ws

                except Exception:
                        print("export.py > createSheetKVPAnaveri Exception!")
                        return None

        def downloadExcel(self, cid_):
                try:
                        wb = Workbook()
                        ws = self.createSheetTanimlar(wb, ModelProfiller, 'Profiller',cid_)
                        ws = self.createSheetTanimlar(wb, ModelBirimler, 'Birimler',cid_)
                        ws = self.createSheetTanimlar(wb, ModelKV, 'Veri Kategorileri',cid_)
                        ws = self.createSheetTanimlar(wb, ModelIslemeAmaclari, 'İşleme Amaçları',cid_)
                        ws = self.createSheetTanimlar(wb, ModelKanallar, 'Kanallar',cid_)
                        ws = self.createSheetTanimlar(wb, ModelSistemler, 'Sistemler',cid_)
                        ws = self.createSheetTanimlar(wb, ModelOrtamlar, 'Ortamlar',cid_)
                        ws = self.createSheetTanimlar(wb, ModelSureler, 'Sureler',cid_)
                        ws = self.createSheetTanimlar(wb, ModelKurumlar, 'Kurumlar',cid_)
                        ws = self.createSheetTanimlar(wb, ModelDayanaklar, 'Dayanaklar',cid_)
                        ws = self.createSheetTanimlar(wb, ModelPaylasimAmaclari, 'PaylasimAmaclari',cid_)
                        ws = self.createSheetTanimlar(wb, ModelPaylasimSekilleri, 'PaylasimSekilleri',cid_)
                        ws = self.createSheetTanimlar(wb, ModelUlkeler, 'Ulkeler',cid_)
                        ws = self.createSheetKVProfil(wb, cid_)
                        ws = self.createSheetKVPaylasim(wb, cid_)
                        ws = self.createSheetKVAnaveri(wb, cid_)

                        # Make header bold
                        # for cell in ws["1:1"]:
                        #     cell.font = Font(bold=True)

                        out = BytesIO()
                        wb.save(out)
                        out.seek(0)

                        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                attachment_filename='gfox.xlsx', as_attachment=True)

                except Exception:
                        return None


def downloadExcel(cid):
    cc = Export()
    return cc.downloadExcel(cid)

